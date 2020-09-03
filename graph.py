


import pandas as pd
import statsmodels.api as sm
from statsmodels.formula.api import poisson
from statsmodels.formula.api import negativebinomial
import numpy as np
import matplotlib.pyplot as plt
import math
from patsy import dmatrices
import openpyxl
# reading data
data_crime = pd.read_excel(r'C:\Users\digiovanniyani\Desktop\NY_CRIMES.xlsx') #, names = ['MH_MURD', 'MH_RAPE', 'MH_ROBB', 'MH_ASSA', 'MH_BURG', 'MH_LARC'])#, delim_whitespace=True, header=0)


#organize data from precinct to borough

crime_file_location=(r"C:\Users\digiovanniyani\Downloads\seven-major-felony-offenses-by-precinct-2000-2019.xls")#USERS
df_crime = pd.read_excel(crime_file_location)

def total_crimes_manhattan(start, end, step): # iterates through the document, step needed to select specific crimes
    list_1=[]
    for i in range(start,end,step):
        list_1.append((df_crime.iloc[i,2:22]))#.values.tolist())
    return(list_1)


a,b,c= 593,625,8

#9,177,8 TOTAL CRIMES MANHATTAN  185,281,8 BX   2 178 8 murder manhattan

list_0=(((total_crimes_manhattan(a,b,c)[0]).values.tolist()))
list_1=(((total_crimes_manhattan(a,b,c)[1]).values.tolist()))
list_2=(((total_crimes_manhattan(a,b,c)[2]).values.tolist()))
list_3=(((total_crimes_manhattan(a,b,c)[3]).values.tolist()))

"""list_4=(((total_crimes_manhattan(a,b,c)[4]).values.tolist()))
list_5=(((total_crimes_manhattan(a,b,c)[5]).values.tolist()))
list_6=(((total_crimes_manhattan(a,b,c)[6]).values.tolist()))
list_7=(((total_crimes_manhattan(a,b,c)[7]).values.tolist()))
list_8=(((total_crimes_manhattan(a,b,c)[8]).values.tolist()))
list_9=(((total_crimes_manhattan(a,b,c)[9]).values.tolist()))
list_10=(((total_crimes_manhattan(a,b,c)[10]).values.tolist()))
list_11=(((total_crimes_manhattan(a,b,c)[11]).values.tolist()))
list_12=(((total_crimes_manhattan(a,b,c)[12]).values.tolist()))
list_13=(((total_crimes_manhattan(a,b,c)[13]).values.tolist()))
list_14=(((total_crimes_manhattan(a,b,c)[14]).values.tolist()))
list_15=(((total_crimes_manhattan(a,b,c)[15]).values.tolist()))
list_16=(((total_crimes_manhattan(a,b,c)[16]).values.tolist()))
list_17=(((total_crimes_manhattan(a,b,c)[17]).values.tolist()))
list_18=(((total_crimes_manhattan(a,b,c)[18]).values.tolist()))
list_19=(((total_crimes_manhattan(a,b,c)[19]).values.tolist()))
list_20=(((total_crimes_manhattan(a,b,c)[20]).values.tolist()))
list_21=(((total_crimes_manhattan(a,b,c)[21]).values.tolist()))
list_22=(((total_crimes_manhattan(a,b,c)[22]).values.tolist()))
list_23=(((total_crimes_manhattan(a,b,c)[23]).values.tolist()))
list_24=(((total_crimes_manhattan(9,185,8)[24]).values.tolist()))
list_25=(((total_crimes_manhattan(9,185,8)[25]).values.tolist()))
list_26=(((total_crimes_manhattan(9,185,8)[26]).values.tolist()))
list_27=(((total_crimes_manhattan(9,185,8)[27]).values.tolist()))
list_28=(((total_crimes_manhattan(9,185,8)[28]).values.tolist()))
list_29=(((total_crimes_manhattan(9,185,8)[29]).values.tolist()))
list_30=(((total_crimes_manhattan(9,185,8)[30]).values.tolist()))
list_31=(((total_crimes_manhattan(9,185,8)[31]).values.tolist()))
list_32=(((total_crimes_manhattan(9,185,8)[32]).values.tolist()))
list_33=(((total_crimes_manhattan(9,185,8)[33]).values.tolist()))"""

print(list_0,list_1,list_2 ,list_3,)

zipped_list = zip(list_0,list_1,list_2 ,list_3)#,list_4,list_5, list_6,list_7,
    #  list_8,list_9,list_10,list_11, list_12,list_13,list_14, list_15)#,list_15, list_16,  list_17, list_18, list_19, list_20, list_21, list_22)
                  #, list_4,list_5, list_6,list_7,list_8, list_9,list_10,list_11,list_12,list_13,list_14,list_15, list_16,  list_17, list_18, list_19, list_20)

number_of_total_crimes_in_manhattan=[sum(i) for i in zipped_list]

print(number_of_total_crimes_in_manhattan)
print(len(number_of_total_crimes_in_manhattan))
#a=len(number_of_total_crimes_in_manhattan)
#print(list_0)"""

wbkName = 'ex.xlsx'
wb = openpyxl.load_workbook(wbkName)
COL=40
ws=wb["Sheet 1"]
for i in range(1,len(number_of_total_crimes_in_manhattan)):
    wcell1 = ws.cell(i, COL)
    wcell1.value =number_of_total_crimes_in_manhattan[i]

ws.move_range("AN1:AN19", rows=2, cols=0)    # Move column to make space for title and first entry
ws.cell(row=1,column=COL).value='SI_TOT'
ws.cell(row=2,column=COL).value=number_of_total_crimes_in_manhattan[0]
wb.save('ex.xlsx')





